"""
AI Akıllı Yemekhane Asistan Sistemi — PDF Rapor Modülü
══════════════════════════════════════════════════════
Haftalık ve aylık raporları PDF olarak oluşturur.

reportlab kullanarak profesyonel görünümlü PDF raporlar üretir:
  - Haftalık özet: genel istatistikler, en beğenilen/az beğenilen yemekler,
    israf skoru, trend yönü
  - Aylık detaylı rapor: günlük detaylar, kategori bazlı analizler

NOT: Windows Arial fontu kullanarak Türkçe karakter desteği sağlar.
"""

import os
import io
import re
from datetime import date, datetime, timedelta

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether, Image,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from sqlalchemy import func as sqla_func
from models import SessionLocal, MenuPuanlama, UretimLog, Yemek


# ═══════════════════════════════════════════════════════════════════
# FONT KAYDI — Türkçe karakter desteği
# ═══════════════════════════════════════════════════════════════════

_FONT_NAME = "Arial"
_FONT_REGISTERED = False


def _register_font():
    """Windows Arial fontunu kaydet (Türkçe karakter desteği için)."""
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return

    # Windows font yolları
    font_paths = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
        r"C:\Windows\Fonts\arialuni.ttf",
    ]

    for path in font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("Arial", path))
                # Bold variant
                bold_paths = [
                    r"C:\Windows\Fonts\arialbd.ttf",
                    r"C:\Windows\Fonts\Arialbd.ttf",
                ]
                for bp in bold_paths:
                    if os.path.exists(bp):
                        pdfmetrics.registerFont(TTFont("Arial-Bold", bp))
                        break
                _FONT_REGISTERED = True
                print(f"✅ PDF font kaydedildi: {path}")
                return
            except Exception as e:
                print(f"⚠️ Font kayıt hatası: {e}")

    # Fallback — DejaVuSans (Linux/Mac)
    try:
        import subprocess
        result = subprocess.run(["fc-list", ":family=DejaVu Sans", "file"],
                                capture_output=True, text=True, timeout=5)
        if result.stdout:
            path = result.stdout.strip().split(":")[0]
            pdfmetrics.registerFont(TTFont("Arial", path))
            _FONT_REGISTERED = True
            return
    except Exception:
        pass

    print("⚠️ Unicode font bulunamadı. Varsayılan Helvetica kullanılacak (Türkçe sorunlu).")


def _strip_emoji(text: str) -> str:
    """Emoji karakterlerini metinden çıkarır."""
    # Emoji Unicode aralıkları
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # Yüz ifadeleri
        "\U0001F300-\U0001F5FF"  # Semboller & Resimler
        "\U0001F680-\U0001F6FF"  # Ulaşım & Harita
        "\U0001F1E0-\U0001F1FF"  # Bayraklar
        "\U00002702-\U000027B0"
        "\U000024C2-\U0001F251"
        "\U0001f926-\U0001f937"
        "\U00010000-\U0010ffff"
        "\u200d"
        "\u2640-\u2642"
        "\u2600-\u2B55"
        "\u23cf"
        "\u23e9"
        "\u231a"
        "\ufe0f"
        "\u3030"
        "]+",
        flags=re.UNICODE,
    )
    return emoji_pattern.sub("", text).strip()


# ═══════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════

def _get_styles():
    """PDF stilleri oluşturur (Arial fontu ile)."""
    _register_font()

    styles = getSampleStyleSheet()
    font = "Arial" if _FONT_REGISTERED else "Helvetica"
    font_bold = "Arial-Bold" if _FONT_REGISTERED else "Helvetica-Bold"

    styles.add(ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName=font_bold,
        fontSize=22,
        textColor=colors.HexColor("#1B4F72"),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName=font,
        fontSize=11,
        textColor=colors.HexColor("#5D6D7E"),
        alignment=TA_CENTER,
        spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontName=font_bold,
        fontSize=14,
        textColor=colors.HexColor("#2E86C1"),
        spaceBefore=16,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        "BodyText2",
        parent=styles["Normal"],
        fontName=font,
        fontSize=9,
    ))
    styles.add(ParagraphStyle(
        "SmallText",
        parent=styles["Normal"],
        fontName=font,
        fontSize=8,
        textColor=colors.HexColor("#8892A4"),
        alignment=TA_CENTER,
    ))
    return styles


def _make_table(data, col_widths=None):
    """Standart tablo stili ile tablo oluşturur."""
    font = "Arial" if _FONT_REGISTERED else "Helvetica"
    font_bold = "Arial-Bold" if _FONT_REGISTERED else "Helvetica-Bold"

    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86C1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("FONTNAME", (0, 1), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D5D8DC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.HexColor("#F8F9FA"), colors.white
        ]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def _get_rating_stats(db, start_date, end_date):
    """Belirli tarih aralığındaki puanlama istatistiklerini döndürür."""
    query = db.query(
        MenuPuanlama.yemek_adi,
        MenuPuanlama.kategori,
        sqla_func.avg(MenuPuanlama.puan).label("ortalama"),
        sqla_func.count(MenuPuanlama.id).label("toplam_oy"),
    ).filter(
        MenuPuanlama.tarih >= start_date,
        MenuPuanlama.tarih <= end_date,
    ).group_by(
        MenuPuanlama.yemek_adi
    ).order_by(
        sqla_func.avg(MenuPuanlama.puan).desc()
    ).all()

    return query


def _get_category_stats(db, start_date, end_date):
    """Kategori bazlı istatistikleri döndürür."""
    query = db.query(
        MenuPuanlama.kategori,
        sqla_func.avg(MenuPuanlama.puan).label("ortalama"),
        sqla_func.count(MenuPuanlama.id).label("toplam_oy"),
    ).filter(
        MenuPuanlama.tarih >= start_date,
        MenuPuanlama.tarih <= end_date,
    ).group_by(
        MenuPuanlama.kategori
    ).all()

    return query


def _get_daily_stats(db, start_date, end_date):
    """Günlük ortalama puanları döndürür."""
    query = db.query(
        MenuPuanlama.tarih,
        sqla_func.avg(MenuPuanlama.puan).label("ortalama"),
        sqla_func.count(MenuPuanlama.id).label("toplam_oy"),
    ).filter(
        MenuPuanlama.tarih >= start_date,
        MenuPuanlama.tarih <= end_date,
    ).group_by(
        MenuPuanlama.tarih
    ).order_by(
        MenuPuanlama.tarih
    ).all()

    return query


# ═══════════════════════════════════════════════════════════════════
# GRAFİK ÜRETİM FONKSİYONLARI (matplotlib → PNG → ReportLab Image)
# ═══════════════════════════════════════════════════════════════════

CHART_COLORS = ["#2E86C1", "#E74C3C", "#27AE60", "#F39C12", "#8E44AD", "#1ABC9C"]


def _chart_daily_trend(daily_stats) -> io.BytesIO | None:
    """Günlük puan trend çizgi grafiği."""
    if not daily_stats or len(daily_stats) < 2:
        return None

    try:
        tarihler = []
        puanlar = []
        for d in daily_stats:
            t = d.tarih if hasattr(d.tarih, "strftime") else datetime.strptime(str(d.tarih), "%Y-%m-%d").date()
            tarihler.append(t)
            puanlar.append(float(d.ortalama))

        fig, ax = plt.subplots(figsize=(6, 2.8))
        ax.plot(tarihler, puanlar, color=CHART_COLORS[0], linewidth=2, marker="o", markersize=4)
        ax.fill_between(tarihler, puanlar, alpha=0.15, color=CHART_COLORS[0])
        ax.set_ylim(0, 5.5)
        ax.set_ylabel("Ort. Puan", fontsize=8)
        ax.set_title("Gunluk Puan Trendi", fontsize=10, fontweight="bold")
        ax.tick_params(axis="both", labelsize=7)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        ax.grid(axis="y", alpha=0.3)
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150)
        plt.close(fig)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _chart_category_avg(cat_stats) -> io.BytesIO | None:
    """Kategori bazlı ortalama puan bar grafiği."""
    if not cat_stats:
        return None

    KAT_LABELS = {
        "corba": "Corba", "ana_yemek": "Ana Yemek",
        "yan_yemek": "Yan Yemek", "tatli": "Tatli", "salata": "Salata",
    }

    try:
        labels = [KAT_LABELS.get(c.kategori, c.kategori or "-") for c in cat_stats]
        values = [float(c.ortalama) for c in cat_stats]

        fig, ax = plt.subplots(figsize=(6, 2.8))
        bars = ax.bar(labels, values, color=CHART_COLORS[:len(labels)], edgecolor="white", width=0.6)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                    f"{val:.1f}", ha="center", va="bottom", fontsize=8, fontweight="bold")
        ax.set_ylim(0, 5.5)
        ax.set_ylabel("Ort. Puan", fontsize=8)
        ax.set_title("Kategori Bazli Ortalama Puanlar", fontsize=10, fontweight="bold")
        ax.tick_params(axis="both", labelsize=7)
        ax.grid(axis="y", alpha=0.3)
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150)
        plt.close(fig)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _chart_waste_trend(db, start_date, end_date) -> io.BytesIO | None:
    """Haftalık israf trend bar grafiği."""
    try:
        haftalik = []
        current = start_date
        while current <= end_date:
            week_end = min(current + timedelta(days=6), end_date)
            row = db.query(
                sqla_func.avg(UretimLog.israf_orani).label("ort"),
            ).filter(
                UretimLog.tarih >= current,
                UretimLog.tarih <= week_end,
            ).first()
            if row and row.ort is not None:
                haftalik.append((f"{current.strftime('%d.%m')}-{week_end.strftime('%d.%m')}", float(row.ort)))
            current = week_end + timedelta(days=1)

        if len(haftalik) < 2:
            return None

        labels, values = zip(*haftalik)
        colors_list = ["#E74C3C" if v >= 40 else "#F39C12" if v >= 25 else "#27AE60" for v in values]

        fig, ax = plt.subplots(figsize=(6, 2.8))
        bars = ax.bar(labels, values, color=colors_list, edgecolor="white", width=0.6)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                    f"%{val:.0f}", ha="center", va="bottom", fontsize=7, fontweight="bold")
        ax.set_ylabel("Israf Orani (%)", fontsize=8)
        ax.set_title("Haftalik Israf Trendi", fontsize=10, fontweight="bold")
        ax.tick_params(axis="both", labelsize=7)
        ax.grid(axis="y", alpha=0.3)
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150)
        plt.close(fig)
        buf.seek(0)
        return buf
    except Exception:
        return None


# ── Varsayılan kategorilere göre maliyet (TL/porsiyon) ────────────
_VARSAYILAN_MALIYET = {
    "corba": 15.0, "ana_yemek": 35.0, "yan_yemek": 10.0,
    "tatli": 20.0, "salata": 12.0, "icecek": 8.0,
}


def _get_maliyet_map(db) -> dict:
    """DB'deki yemek birim maliyetlerini {ad_lower: maliyet} dict olarak döndürür."""
    yemekler = db.query(Yemek).all()
    m = {}
    for y in yemekler:
        if y.birim_maliyet is not None:
            m[y.ad.lower()] = y.birim_maliyet
    return m


def _birim_maliyet(maliyet_map, yemek_adi, kategori) -> float:
    """Yemek için birim maliyeti döndürür; DB'de yoksa varsayılan kullanır."""
    return maliyet_map.get(yemek_adi.lower(), _VARSAYILAN_MALIYET.get(kategori, 20.0))


def _chart_cost_pie(db, start_date, end_date, porsiyon_maliyet=None) -> io.BytesIO | None:
    """Maliyet dağılımı pasta grafiği (tüketilen vs israf — gerçek birim maliyet)."""
    try:
        maliyet_map = _get_maliyet_map(db)

        logs = db.query(UretimLog).filter(
            UretimLog.tarih >= start_date,
            UretimLog.tarih <= end_date,
        ).all()

        if not logs:
            return None

        tuketim_maliyet = 0.0
        israf_maliyet = 0.0
        for log in logs:
            birim = _birim_maliyet(maliyet_map, log.yemek_adi, log.kategori)
            uretilen = log.uretilen_porsiyon or 0
            kalan = log.kalan_porsiyon or 0
            tuketilen = max(uretilen - kalan, 0)
            tuketim_maliyet += tuketilen * birim
            israf_maliyet += kalan * birim

        if (tuketim_maliyet + israf_maliyet) == 0:
            return None

        fig, ax = plt.subplots(figsize=(4, 2.8))
        sizes = [tuketim_maliyet, israf_maliyet]
        labels_list = [
            f"Tuketilen\n{tuketim_maliyet:,.0f} TL",
            f"Israf\n{israf_maliyet:,.0f} TL",
        ]
        colors_pie = ["#27AE60", "#E74C3C"]
        wedges, texts, autotexts = ax.pie(
            sizes, labels=labels_list, colors=colors_pie, autopct="%1.1f%%",
            startangle=90, textprops={"fontsize": 8},
        )
        for t in autotexts:
            t.set_fontsize(8)
            t.set_fontweight("bold")
        ax.set_title("Maliyet Dagilimi", fontsize=10, fontweight="bold")
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150)
        plt.close(fig)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _build_maliyet_section(db, start_date, end_date, styles) -> list:
    """PDF'e eklenecek maliyet analiz bölümü elemanlarını döndürür."""
    elements = []
    try:
        maliyet_map = _get_maliyet_map(db)

        logs = db.query(UretimLog).filter(
            UretimLog.tarih >= start_date,
            UretimLog.tarih <= end_date,
        ).all()

        if not logs:
            return elements

        yemek_toplam = {}  # {yemek_adi: {kalan, kayip_tl, kategori}}
        genel_kayip = 0.0

        for log in logs:
            birim = _birim_maliyet(maliyet_map, log.yemek_adi, log.kategori)
            kalan = log.kalan_porsiyon or 0
            kayip = kalan * birim
            genel_kayip += kayip

            key = log.yemek_adi
            if key not in yemek_toplam:
                yemek_toplam[key] = {"kalan": 0, "kayip_tl": 0.0, "kategori": log.kategori, "birim": birim}
            yemek_toplam[key]["kalan"] += kalan
            yemek_toplam[key]["kayip_tl"] += kayip

        elements.append(Paragraph("Israf Maliyet Analizi", styles["SectionTitle"]))
        elements.append(Paragraph(
            f"Toplam israf maliyeti: {genel_kayip:,.0f} TL",
            styles["BodyText2"],
        ))
        elements.append(Spacer(1, 8))

        # En maliyetli israf yemekler tablosu
        sirali = sorted(yemek_toplam.items(), key=lambda x: x[1]["kayip_tl"], reverse=True)
        tablo_data = [["Yemek", "Kategori", "Kalan Prs.", "Birim (TL)", "Kayip (TL)"]]
        for yemek_adi, v in sirali[:10]:
            tablo_data.append([
                _strip_emoji(yemek_adi),
                _strip_emoji(v["kategori"] or "-"),
                str(int(v["kalan"])),
                f"{v['birim']:.0f}",
                f"{v['kayip_tl']:,.0f}",
            ])
        elements.append(_make_table(tablo_data, col_widths=[4.5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]))

    except Exception:
        pass
    return elements


def _build_donem_karsilastirma_section(db, bu_bas, bu_son, gun_sayisi, styles) -> list:
    """PDF'e eklenecek karşılaştırmalı dönem raporu elemanlarını döndürür."""
    elements = []
    try:
        gecen_son = bu_bas - timedelta(days=1)
        gecen_bas = gecen_son - timedelta(days=gun_sayisi - 1)

        donem_label = "Hafta" if gun_sayisi <= 7 else "Ay"

        def _donem_stats(bas, son):
            israf = db.query(
                sqla_func.avg(UretimLog.israf_orani).label("ort_israf"),
                sqla_func.sum(UretimLog.uretilen_porsiyon).label("uretilen"),
                sqla_func.sum(UretimLog.kalan_porsiyon).label("kalan"),
                sqla_func.count(UretimLog.id).label("kayit"),
            ).filter(
                UretimLog.tarih >= bas, UretimLog.tarih <= son,
                UretimLog.israf_orani.isnot(None),
            ).first()

            puan = db.query(
                sqla_func.avg(MenuPuanlama.puan).label("ort_puan"),
                sqla_func.count(MenuPuanlama.id).label("oy"),
            ).filter(
                MenuPuanlama.tarih >= bas, MenuPuanlama.tarih <= son,
            ).first()

            ort_israf = round(float(israf.ort_israf or 0), 1)
            kalan = int(israf.kalan or 0)
            uretilen = int(israf.uretilen or 0)

            # Maliyet
            maliyet_map = _get_maliyet_map(db)
            maliyet_kayip = 0.0
            logs = db.query(UretimLog).filter(
                UretimLog.tarih >= bas, UretimLog.tarih <= son,
            ).all()
            for log in logs:
                birim = _birim_maliyet(maliyet_map, log.yemek_adi, log.kategori)
                maliyet_kayip += (log.kalan_porsiyon or 0) * birim

            return {
                "ort_israf": ort_israf,
                "uretilen": uretilen,
                "kalan": kalan,
                "kayit": int(israf.kayit or 0),
                "ort_puan": round(float(puan.ort_puan or 0), 1) if puan.ort_puan else 0,
                "oy": int(puan.oy or 0),
                "maliyet_tl": round(maliyet_kayip, 0),
            }

        bu = _donem_stats(bu_bas, bu_son)
        gecen = _donem_stats(gecen_bas, gecen_son)

        # İyileşme
        def _iyilesme(eski, yeni):
            return round((eski - yeni) / eski * 100, 1) if eski != 0 else 0

        israf_fark = round(bu["ort_israf"] - gecen["ort_israf"], 1)
        israf_pct = _iyilesme(gecen["ort_israf"], bu["ort_israf"])
        maliyet_fark = round(gecen["maliyet_tl"] - bu["maliyet_tl"], 0)
        puan_fark = round(bu["ort_puan"] - gecen["ort_puan"], 2)

        elements.append(Paragraph(
            f"Karsilastirmali Donem Analizi (Bu {donem_label} vs Gecen {donem_label})",
            styles["SectionTitle"],
        ))

        # Durum özeti
        if israf_fark < -1:
            durum = f"Olumlu: Israf %{abs(israf_fark)} azaldi ({israf_pct:+.1f}%). "
        elif israf_fark > 1:
            durum = f"Olumsuz: Israf %{israf_fark} artti. Onlem onerisi gerekli. "
        else:
            durum = "Israf orani stabil. "

        if maliyet_fark > 0:
            durum += f"Maliyet tasarrufu: {maliyet_fark:,.0f} TL. "
        elif maliyet_fark < 0:
            durum += f"Maliyet artisi: {abs(maliyet_fark):,.0f} TL. "

        if puan_fark > 0:
            durum += f"Puanlar {puan_fark:+.2f} iyilesti."
        elif puan_fark < 0:
            durum += f"Puanlar {puan_fark:+.2f} dustu."

        elements.append(Paragraph(durum, styles["BodyText2"]))
        elements.append(Spacer(1, 8))

        # Tablo
        tablo_data = [
            ["Metrik", f"Gecen {donem_label}", f"Bu {donem_label}", "Degisim"],
            ["Ort. Israf (%)", f"%{gecen['ort_israf']}", f"%{bu['ort_israf']}", f"{'+' if israf_fark > 0 else ''}{israf_fark}"],
            ["Israf Maliyeti (TL)", f"{gecen['maliyet_tl']:,.0f}", f"{bu['maliyet_tl']:,.0f}", f"{'+' if maliyet_fark < 0 else '-'}{abs(maliyet_fark):,.0f}"],
            ["Ort. Puan", f"{gecen['ort_puan']}/5", f"{bu['ort_puan']}/5", f"{'+' if puan_fark > 0 else ''}{puan_fark}"],
            ["Toplam Oy", str(gecen["oy"]), str(bu["oy"]), f"{'+' if bu['oy'] > gecen['oy'] else ''}{bu['oy'] - gecen['oy']}"],
            ["Uretilen Prs.", str(gecen["uretilen"]), str(bu["uretilen"]), f"{'+' if bu['uretilen'] > gecen['uretilen'] else ''}{bu['uretilen'] - gecen['uretilen']}"],
            ["Kalan Prs.", str(gecen["kalan"]), str(bu["kalan"]), f"{'+' if bu['kalan'] > gecen['kalan'] else ''}{bu['kalan'] - gecen['kalan']}"],
        ]
        elements.append(_make_table(tablo_data, col_widths=[4*cm, 3.5*cm, 3.5*cm, 3.5*cm]))

    except Exception:
        pass
    return elements


# ═══════════════════════════════════════════════════════════════════
# HAFTALIK PDF RAPOR
# ═══════════════════════════════════════════════════════════════════

def generate_weekly_pdf(db=None) -> bytes:
    """
    Son 7 günlük haftalık özet raporunu PDF olarak üretir.
    """
    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        today = date.today()
        start = today - timedelta(days=6)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=2 * cm, bottomMargin=2 * cm,
            leftMargin=2.5 * cm, rightMargin=2.5 * cm,
        )
        styles = _get_styles()
        elements = []

        # ─── Başlık ─────────────────────────────────────────────
        elements.append(Paragraph(
            "Akilli Yemekhane - Haftalik Rapor", styles["ReportTitle"]
        ))
        elements.append(Paragraph(
            f"{start.strftime('%d.%m.%Y')} - {today.strftime('%d.%m.%Y')}",
            styles["ReportSubtitle"],
        ))
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=colors.HexColor("#2E86C1"), spaceAfter=16,
        ))

        # ─── Genel İstatistikler ────────────────────────────────
        all_ratings = _get_rating_stats(db, start, today)
        total_votes = sum(r.toplam_oy for r in all_ratings)
        if all_ratings:
            overall_avg = sum(r.ortalama * r.toplam_oy for r in all_ratings) / max(total_votes, 1)
        else:
            overall_avg = 0

        elements.append(Paragraph("Genel Ozet", styles["SectionTitle"]))

        summary_data = [
            ["Metrik", _strip_emoji("Deger")],
            ["Rapor Donemi", f"{start.strftime('%d.%m.%Y')} - {today.strftime('%d.%m.%Y')}"],
            ["Toplam Puanlama Sayisi", str(total_votes)],
            ["Genel Ortalama Puan", f"{overall_avg:.1f} / 5"],
            [_strip_emoji("Degerlendirilen Yemek Sayisi"), str(len(all_ratings))],
        ]
        elements.append(_make_table(summary_data, col_widths=[7 * cm, 8 * cm]))
        elements.append(Spacer(1, 16))

        # ─── En Beğenilen 10 Yemek ──────────────────────────────
        elements.append(Paragraph(
            "En Begenilen Yemekler (Top 10)", styles["SectionTitle"]
        ))
        if all_ratings:
            top_data = [["#", "Yemek Adi", "Kategori", "Ort. Puan", "Oy Sayisi"]]
            for i, r in enumerate(all_ratings[:10], 1):
                top_data.append([
                    str(i),
                    _strip_emoji(r.yemek_adi),
                    _strip_emoji(r.kategori or "-"),
                    f"{r.ortalama:.1f}",
                    str(r.toplam_oy),
                ])
            elements.append(_make_table(top_data, col_widths=[1.2*cm, 5*cm, 3*cm, 2.8*cm, 2.5*cm]))
        else:
            elements.append(Paragraph("Veri bulunamadi.", styles["BodyText2"]))
        elements.append(Spacer(1, 16))

        # ─── En Az Beğenilen 5 Yemek ────────────────────────────
        elements.append(Paragraph(
            "En Az Begenilen Yemekler", styles["SectionTitle"]
        ))
        if len(all_ratings) >= 5:
            bottom = list(reversed(all_ratings))[:5]
            bot_data = [["#", "Yemek Adi", "Kategori", "Ort. Puan", "Oy Sayisi"]]
            for i, r in enumerate(bottom, 1):
                bot_data.append([
                    str(i),
                    _strip_emoji(r.yemek_adi),
                    _strip_emoji(r.kategori or "-"),
                    f"{r.ortalama:.1f}",
                    str(r.toplam_oy),
                ])
            elements.append(_make_table(bot_data, col_widths=[1.2*cm, 5*cm, 3*cm, 2.8*cm, 2.5*cm]))
        else:
            elements.append(Paragraph("Yeterli veri yok.", styles["BodyText2"]))
        elements.append(Spacer(1, 16))

        # ─── Kategori Bazlı ─────────────────────────────────────
        cat_stats = _get_category_stats(db, start, today)
        elements.append(Paragraph(
            "Kategori Bazli Ortalamalar", styles["SectionTitle"]
        ))

        KAT_LABELS = {
            "corba": "Corbalar", "ana_yemek": "Ana Yemekler",
            "yan_yemek": "Yan Yemek", "tatli": "Tatlilar",
            "salata": "Salatalar",
        }

        if cat_stats:
            cat_data = [["Kategori", "Ortalama Puan", "Toplam Oy"]]
            for c in cat_stats:
                cat_data.append([
                    KAT_LABELS.get(c.kategori, c.kategori or "-"),
                    f"{c.ortalama:.1f}",
                    str(c.toplam_oy),
                ])
            elements.append(_make_table(cat_data, col_widths=[5 * cm, 5 * cm, 4.5 * cm]))
            # Kategori bar grafiği
            cat_chart = _chart_category_avg(cat_stats)
            if cat_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(cat_chart, width=14 * cm, height=6.5 * cm))
        else:
            elements.append(Paragraph("Veri bulunamadi.", styles["BodyText2"]))
        elements.append(Spacer(1, 16))

        # ─── Günlük Trend ───────────────────────────────────────
        daily_stats = _get_daily_stats(db, start, today)
        elements.append(Paragraph(
            "Gunluk Puan Trendi", styles["SectionTitle"]
        ))

        if daily_stats:
            day_data = [["Tarih", "Ortalama Puan", "Oy Sayisi"]]
            for d in daily_stats:
                tarih_str = d.tarih.strftime("%d.%m.%Y") if hasattr(d.tarih, "strftime") else str(d.tarih)
                day_data.append([
                    tarih_str,
                    f"{d.ortalama:.1f}",
                    str(d.toplam_oy),
                ])
            elements.append(_make_table(day_data, col_widths=[5 * cm, 5 * cm, 4.5 * cm]))
            # Günlük trend çizgi grafiği
            trend_chart = _chart_daily_trend(daily_stats)
            if trend_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(trend_chart, width=14 * cm, height=6.5 * cm))
        else:
            elements.append(Paragraph("Veri bulunamadi.", styles["BodyText2"]))

        # ─── Israf Tahmini ───────────────────────────────────────
        elements.append(Spacer(1, 16))
        elements.append(Paragraph(
            "Israf Tahmini ve Uretim Analizi", styles["SectionTitle"]
        ))

        try:
            uretim_stats = db.query(
                UretimLog.yemek_adi,
                sqla_func.avg(UretimLog.israf_orani).label("ort_israf"),
                sqla_func.avg(UretimLog.tuketim_orani).label("ort_tuketim"),
                sqla_func.sum(UretimLog.uretilen_porsiyon).label("toplam_uretim"),
                sqla_func.count(UretimLog.id).label("kayit"),
            ).filter(
                UretimLog.tarih >= start,
                UretimLog.tarih <= today,
            ).group_by(
                UretimLog.yemek_adi
            ).order_by(
                sqla_func.avg(UretimLog.israf_orani).desc()
            ).limit(10).all()

            if uretim_stats:
                israf_data = [["Yemek", "Ort. Israf %", "Ort. Tuketim %", "Toplam Uretim"]]
                for u in uretim_stats:
                    israf_data.append([
                        _strip_emoji(u.yemek_adi),
                        f"%{u.ort_israf:.1f}",
                        f"%{u.ort_tuketim:.1f}",
                        str(int(u.toplam_uretim or 0)),
                    ])
                elements.append(_make_table(israf_data, col_widths=[5*cm, 3*cm, 3.5*cm, 3*cm]))

                # Genel israf ozeti
                genel_israf = db.query(
                    sqla_func.avg(UretimLog.israf_orani)
                ).filter(
                    UretimLog.tarih >= start,
                    UretimLog.tarih <= today,
                ).scalar()

                if genel_israf is not None:
                    elements.append(Spacer(1, 8))
                    seviye = "KRITIK" if genel_israf >= 50 else "UYARI" if genel_israf >= 30 else "NORMAL"
                    elements.append(Paragraph(
                        f"Haftalik genel israf orani: %{genel_israf:.1f} — Seviye: {seviye}",
                        styles["BodyText2"],
                    ))
            else:
                elements.append(Paragraph("Uretim log verisi bulunamadi.", styles["BodyText2"]))

            # İsraf trend grafiği
            waste_chart = _chart_waste_trend(db, start, today)
            if waste_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(waste_chart, width=14 * cm, height=6.5 * cm))

            # Maliyet pasta grafiği
            cost_chart = _chart_cost_pie(db, start, today)
            if cost_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(cost_chart, width=10 * cm, height=7 * cm))

        except Exception:
            elements.append(Paragraph("Israf verisi yuklenemedi.", styles["BodyText2"]))

        # ─── Maliyet Analizi Bölümü ───────────────────────────
        elements.append(Spacer(1, 16))
        maliyet_elems = _build_maliyet_section(db, start, today, styles)
        elements.extend(maliyet_elems)

        # ─── Karşılaştırmalı Dönem Raporu (2C) ───────────────
        elements.append(Spacer(1, 16))
        donem_elems = _build_donem_karsilastirma_section(db, start, today, 7, styles)
        elements.extend(donem_elems)

        # ─── Footer ─────────────────────────────────────────────
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor("#D5D8DC"), spaceAfter=8,
        ))
        elements.append(Paragraph(
            f"Bu rapor AI Akilli Yemekhane Asistan Sistemi tarafindan otomatik olusturulmustur. "
            f"Olusturma tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["SmallText"],
        ))

        doc.build(elements)
        return buf.getvalue()

    finally:
        if close_db:
            db.close()


# ═══════════════════════════════════════════════════════════════════
# AYLIK PDF RAPOR
# ═══════════════════════════════════════════════════════════════════

def generate_monthly_pdf(db=None) -> bytes:
    """
    Son 30 günlük aylık detaylı raporunu PDF olarak üretir.
    """
    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        today = date.today()
        start = today - timedelta(days=29)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=2 * cm, bottomMargin=2 * cm,
            leftMargin=2.5 * cm, rightMargin=2.5 * cm,
        )
        styles = _get_styles()
        elements = []

        # ─── Başlık ─────────────────────────────────────────────
        elements.append(Paragraph(
            "Akilli Yemekhane - Aylik Detayli Rapor", styles["ReportTitle"]
        ))
        elements.append(Paragraph(
            f"{start.strftime('%d.%m.%Y')} - {today.strftime('%d.%m.%Y')} (30 Gun)",
            styles["ReportSubtitle"],
        ))
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=colors.HexColor("#2E86C1"), spaceAfter=16,
        ))

        # ─── Genel İstatistikler ────────────────────────────────
        all_ratings = _get_rating_stats(db, start, today)
        total_votes = sum(r.toplam_oy for r in all_ratings)
        if all_ratings:
            overall_avg = sum(r.ortalama * r.toplam_oy for r in all_ratings) / max(total_votes, 1)
        else:
            overall_avg = 0

        elements.append(Paragraph("30 Gunluk Genel Ozet", styles["SectionTitle"]))

        summary_data = [
            ["Metrik", "Deger"],
            ["Rapor Donemi", f"{start.strftime('%d.%m.%Y')} - {today.strftime('%d.%m.%Y')}"],
            ["Toplam Puanlama", str(total_votes)],
            ["Genel Ortalama", f"{overall_avg:.1f} / 5"],
            ["Farkli Yemek", str(len(all_ratings))],
        ]
        elements.append(_make_table(summary_data, col_widths=[7 * cm, 8 * cm]))
        elements.append(Spacer(1, 16))

        # ─── Tüm Yemekler Detaylı Tablo ─────────────────────────
        elements.append(Paragraph(
            "Tum Yemekler - Detayli Istatistik", styles["SectionTitle"]
        ))
        if all_ratings:
            full_data = [["#", "Yemek", "Kategori", "Ort.", "Oy", "Seviye"]]
            for i, r in enumerate(all_ratings, 1):
                if r.ortalama >= 4.0:
                    level = "Cok Iyi"
                elif r.ortalama >= 3.0:
                    level = "Iyi"
                elif r.ortalama >= 2.0:
                    level = "Orta"
                else:
                    level = "Kotu"
                full_data.append([
                    str(i), _strip_emoji(r.yemek_adi),
                    _strip_emoji(r.kategori or "-"),
                    f"{r.ortalama:.1f}", str(r.toplam_oy), level,
                ])
            elements.append(_make_table(
                full_data,
                col_widths=[1*cm, 4.5*cm, 2.8*cm, 1.8*cm, 1.8*cm, 2.6*cm],
            ))
        else:
            elements.append(Paragraph("Veri bulunamadi.", styles["BodyText2"]))
        elements.append(Spacer(1, 16))

        # ─── Kategori Bazlı ─────────────────────────────────────
        cat_stats = _get_category_stats(db, start, today)
        elements.append(Paragraph("Kategori Performansi", styles["SectionTitle"]))

        KAT_LABELS = {
            "corba": "Corbalar", "ana_yemek": "Ana Yemekler",
            "yan_yemek": "Yan Yemek", "tatli": "Tatlilar",
            "salata": "Salatalar",
        }
        if cat_stats:
            cat_data = [["Kategori", "Ortalama", "Oy"]]
            for c in cat_stats:
                cat_data.append([
                    KAT_LABELS.get(c.kategori, c.kategori or "-"),
                    f"{c.ortalama:.1f}", str(c.toplam_oy),
                ])
            elements.append(_make_table(cat_data, col_widths=[5*cm, 5*cm, 4.5*cm]))
            # Kategori bar grafiği
            cat_chart = _chart_category_avg(cat_stats)
            if cat_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(cat_chart, width=14 * cm, height=6.5 * cm))

        # ─── Günlük Detay ───────────────────────────────────────
        daily_stats = _get_daily_stats(db, start, today)
        elements.append(Spacer(1, 16))
        elements.append(Paragraph(
            "Gunluk Puan Trendi (30 Gun)", styles["SectionTitle"]
        ))

        if daily_stats:
            day_data = [["Tarih", "Ort. Puan", "Oy", "Durum"]]
            for d in daily_stats:
                tarih_str = d.tarih.strftime("%d.%m.%Y") if hasattr(d.tarih, "strftime") else str(d.tarih)
                if d.ortalama >= 3.5:
                    durum = "Iyi"
                elif d.ortalama >= 2.5:
                    durum = "Orta"
                else:
                    durum = "Kotu"
                day_data.append([tarih_str, f"{d.ortalama:.1f}", str(d.toplam_oy), durum])
            elements.append(_make_table(day_data, col_widths=[4*cm, 3.5*cm, 3*cm, 4*cm]))
            # Günlük trend çizgi grafiği
            trend_chart = _chart_daily_trend(daily_stats)
            if trend_chart:
                elements.append(Spacer(1, 8))
                elements.append(Image(trend_chart, width=14 * cm, height=6.5 * cm))

        # ─── İsraf & Maliyet Grafikleri ──────────────────────────
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("Israf ve Maliyet Analizi", styles["SectionTitle"]))

        waste_chart = _chart_waste_trend(db, start, today)
        if waste_chart:
            elements.append(Image(waste_chart, width=14 * cm, height=6.5 * cm))
            elements.append(Spacer(1, 8))

        cost_chart = _chart_cost_pie(db, start, today)
        if cost_chart:
            elements.append(Image(cost_chart, width=10 * cm, height=7 * cm))

        # ─── Detaylı Maliyet Analizi Bölümü ──────────────────────
        elements.append(Spacer(1, 16))
        maliyet_elems = _build_maliyet_section(db, start, today, styles)
        elements.extend(maliyet_elems)

        # ─── Karşılaştırmalı Dönem Analizi (Bu ay vs Geçen ay) ──
        elements.append(Spacer(1, 16))
        donem_elems = _build_donem_karsilastirma_section(db, start, today, 30, styles)
        elements.extend(donem_elems)

        # ─── Model Performans Ozeti ─────────────────────────────
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("AI Model Performans Ozeti", styles["SectionTitle"]))

        try:
            from modules.model_tracker import get_latest_metrics
            latest = get_latest_metrics()
            if latest:
                model_data = [["Model", "Metrik", "Deger", "Tarih"]]
                for entry in latest:
                    for mkey, mval in entry.get("metrics", {}).items():
                        model_data.append([
                            entry.get("model", "-"),
                            mkey,
                            f"{mval:.4f}" if isinstance(mval, float) else str(mval),
                            entry.get("trained_at", "-")[:10],
                        ])
                if len(model_data) > 1:
                    elements.append(_make_table(model_data, col_widths=[3.5*cm, 3.5*cm, 3.5*cm, 4*cm]))
                else:
                    elements.append(Paragraph("Henuz model metrigi kaydedilmemis.", styles["BodyText2"]))
            else:
                elements.append(Paragraph("Model tracker verisi bulunamadi.", styles["BodyText2"]))
        except Exception:
            elements.append(Paragraph("Model metrikleri yuklenemedi.", styles["BodyText2"]))

        # ─── Footer ─────────────────────────────────────────────
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor("#D5D8DC"), spaceAfter=8,
        ))
        elements.append(Paragraph(
            f"Bu rapor AI Akilli Yemekhane Asistan Sistemi tarafindan otomatik olusturulmustur. "
            f"Olusturma tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["SmallText"],
        ))

        doc.build(elements)
        return buf.getvalue()

    finally:
        if close_db:
            db.close()


# ═══════════════════════════════════════════════════════════════════
# EXCEL RAPOR EXPORT
# ═══════════════════════════════════════════════════════════════════

def generate_excel_report(db=None, gun: int = 30) -> bytes:
    """
    Son N günlük verileri Excel (.xlsx) olarak export eder.
    Sayfalar: Puanlama Özeti, Günlük Detay, Kategori, İsraf Analizi
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl kurulu değil. pip install openpyxl")

    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        today = date.today()
        start = today - timedelta(days=gun - 1)

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D5D8DC"),
            right=Side(style="thin", color="D5D8DC"),
            top=Side(style="thin", color="D5D8DC"),
            bottom=Side(style="thin", color="D5D8DC"),
        )

        def _style_header(ws, col_count):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def _style_cells(ws, row_count, col_count):
            for r in range(1, row_count + 1):
                for c in range(1, col_count + 1):
                    ws.cell(row=r, column=c).border = thin_border
                    if r > 1:
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

        # ── Sayfa 1: Puanlama Özeti ──
        ws1 = wb.active
        ws1.title = "Puanlama Ozeti"
        ws1.append(["Yemek Adi", "Kategori", "Ortalama Puan", "Toplam Oy"])
        ratings = _get_rating_stats(db, start, today)
        for r in ratings:
            ws1.append([r.yemek_adi, r.kategori or "-", round(float(r.ortalama), 2), r.toplam_oy])
        _style_header(ws1, 4)
        _style_cells(ws1, len(ratings) + 1, 4)
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 15
        ws1.column_dimensions["C"].width = 15
        ws1.column_dimensions["D"].width = 12

        # ── Sayfa 2: Günlük Detay ──
        ws2 = wb.create_sheet("Gunluk Detay")
        ws2.append(["Tarih", "Ortalama Puan", "Oy Sayisi"])
        daily = _get_daily_stats(db, start, today)
        for d in daily:
            tarih_str = d.tarih.strftime("%Y-%m-%d") if hasattr(d.tarih, "strftime") else str(d.tarih)
            ws2.append([tarih_str, round(float(d.ortalama), 2), d.toplam_oy])
        _style_header(ws2, 3)
        _style_cells(ws2, len(daily) + 1, 3)
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 12

        # ── Sayfa 3: Kategori Bazlı ──
        ws3 = wb.create_sheet("Kategori Analizi")
        ws3.append(["Kategori", "Ortalama Puan", "Toplam Oy"])
        cats = _get_category_stats(db, start, today)
        for c in cats:
            ws3.append([c.kategori or "-", round(float(c.ortalama), 2), c.toplam_oy])
        _style_header(ws3, 3)
        _style_cells(ws3, len(cats) + 1, 3)
        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 15
        ws3.column_dimensions["C"].width = 12

        # ── Sayfa 4: İsraf Analizi ──
        ws4 = wb.create_sheet("Israf Analizi")
        ws4.append(["Yemek Adi", "Kategori", "Ort. Israf %", "Ort. Tuketim %", "Toplam Uretim", "Kayit"])
        israf_stats = db.query(
            UretimLog.yemek_adi,
            UretimLog.kategori,
            sqla_func.avg(UretimLog.israf_orani).label("ort_israf"),
            sqla_func.avg(UretimLog.tuketim_orani).label("ort_tuketim"),
            sqla_func.sum(UretimLog.uretilen_porsiyon).label("toplam_uretim"),
            sqla_func.count(UretimLog.id).label("kayit"),
        ).filter(
            UretimLog.tarih >= start, UretimLog.tarih <= today,
        ).group_by(UretimLog.yemek_adi, UretimLog.kategori).order_by(
            sqla_func.avg(UretimLog.israf_orani).desc()
        ).all()
        for u in israf_stats:
            ws4.append([
                u.yemek_adi, u.kategori or "-",
                round(float(u.ort_israf or 0), 1),
                round(float(u.ort_tuketim or 0), 1),
                int(u.toplam_uretim or 0), u.kayit,
            ])
        _style_header(ws4, 6)
        _style_cells(ws4, len(israf_stats) + 1, 6)
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 15

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        return excel_buf.getvalue()

    finally:
        if close_db:
            db.close()
