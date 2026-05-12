# -*- coding: utf-8 -*-
"""Read all monthly xlsx files and dump menu data."""
import openpyxl
import os
import sys
import re
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\cemil\Downloads\GRADUATİON PROJECT\aylık yemek listesi'

for f in sorted(os.listdir(folder)):
    if not f.endswith('.xlsx'):
        continue
    path = os.path.join(folder, f)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print(f'\n{"="*60}')
    print(f'FILE: {f}  (rows={ws.max_row}, cols={ws.max_column})')
    print(f'{"="*60}')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False)):
        vals = []
        for c in row:
            v = c.value
            if v is None:
                vals.append('')
            elif isinstance(v, datetime):
                vals.append(v.strftime('%Y-%m-%d'))
            else:
                vals.append(str(v)[:70])
        if any(v for v in vals):
            print(f'  R{i+1}: {vals}')
