# -*- coding: utf-8 -*-
"""
Tüm aylık Excel yemek listelerini okuyup menu_data.csv'yi yeniden oluşturur.

Excel yapısı (her ay):
  R1: Üniversite adı
  R2: Birim adı
  R3: Ay başlığı
  R4: [tarih] [tarih] [tarih] [tarih] [tarih]   <- Hafta 1
  R5: çorba    çorba   çorba   çorba   çorba
  R6: ana      ana     ana     ana     ana
  R7: pilav    pilav   pilav   pilav   pilav
  R8: tatlı    tatlı   tatlı   tatlı   tatlı
  R9: [tarih] [tarih] [tarih] [tarih] [tarih]   <- Hafta 2
  R10-R13: yemekler
  ...
"""
import openpyxl
import os
import sys
import re
import csv
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r'c:\Users\cemil\Downloads\GRADUATİON PROJECT\aylara göre yemek listeleri'
OUTPUT = r'c:\Users\cemil\Downloads\GRADUATİON PROJECT\ai_yemekhane\data\menu_data.csv'

GUN_MAP = {
    0: 'Pazartesi', 1: 'Salı', 2: 'Çarşamba',
    3: 'Perşembe', 4: 'Cuma', 5: 'Cumartesi', 6: 'Pazar'
}


def clean_meal(text):
    if not text:
        return ''
    text = str(text).strip()
    if not text or text == 'None':
        return ''
    if 'RESMİ TATİL' in text.upper() or 'TATIL' in text.upper():
        return ''
    # "Mercimek Çorba 186 kkal" -> "Mercimek Çorba"
    text = re.sub(r'\s*\d+\s*k?ka[lq]+\s*$', '', text, flags=re.IGNORECASE).strip()
    # Sondaki kalori sayıları: "Et Fajita 371" -> "Et Fajita"
    text = re.sub(r'\s+\d{2,4}\s*$', '', text).strip()
    return text


def classify_4th_item(text):
    """4. satır yemek: tatlı mı salata mı?"""
    if not text:
        return 'tatli', ''
    tl = text.lower()

    TATLI_KW = ['komposto', 'puding', 'helva', 'baklava', 'tatlı', 'kek', 'tiramisu',
                'trileçe', 'triliçe', 'keşkül', 'helvası', 'şekerpare', 'brownie', 'sultan',
                'sütlaç', 'muhallebi', 'ekler', 'kalburabastı', 'irmik', 'kemalpaşa',
                'kıbrıs', 'balbadem', 'meyve suyu', 'sarışınım', 'sevgi', 'islak',
                'incirli', 'hoşaf', 'revani', 'tulumba', 'vezir', 'kabak tatlısı']
    SALATA_KW = ['salata', 'cacık', 'ayran', 'yoğurt', 'turşu', 'haydari',
                 'piyaz', 'tarator', 'borani', 'söğüş', 'coleslow', 'iceberg',
                 'lahana']
    MEYVE_KW = ['meyve', 'portakal', 'elma', 'muz', 'armut', 'mandalina', 'üzüm', 'kavun']

    is_tatli = any(k in tl for k in TATLI_KW)
    is_salata = any(k in tl for k in SALATA_KW)
    is_meyve = any(k in tl for k in MEYVE_KW) and not is_tatli

    if is_tatli:
        return 'tatli', text
    elif is_salata:
        return 'salata', text
    elif is_meyve:
        return 'salata', text  # Meyve = salata/yan ürün
    else:
        return 'salata', text  # Varsayılan


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    results = []
    max_col = ws.max_column
    max_row = ws.max_row

    # TÜM tarih satırlarını bul
    date_rows = []
    for r in range(1, max_row + 1):
        has_date = False
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, datetime):
                has_date = True
                break
        if has_date:
            date_rows.append(r)

    if not date_rows:
        print(f"  UYARI: Tarih satırı bulunamadı: {os.path.basename(filepath)}")
        return []

    # Her tarih satırı için sonraki 4 satırı yemek olarak oku
    for dr in date_rows:
        for col in range(1, max_col + 1):
            date_val = ws.cell(row=dr, column=col).value
            if not isinstance(date_val, datetime):
                continue

            tarih = date_val.date()
            gun = GUN_MAP.get(tarih.weekday(), '')

            # Sonraki 4 satır: çorba, ana yemek, pilav, tatlı/salata
            meal_rows = []
            for mr in range(dr + 1, min(dr + 5, max_row + 1)):
                val = ws.cell(row=mr, column=col).value
                cleaned = clean_meal(val)
                meal_rows.append(cleaned)

            # Pad to 4
            while len(meal_rows) < 4:
                meal_rows.append('')

            corba = meal_rows[0]
            ana_yemek = meal_rows[1]
            pilav = meal_rows[2]
            dorduncu = meal_rows[3]

            # RESMİ TATİL kontrolü
            if not corba and not ana_yemek:
                continue

            # 4. satırı sınıflandır
            kategori, yemek = classify_4th_item(dorduncu)
            tatli = yemek if kategori == 'tatli' else ''
            salata = yemek if kategori == 'salata' else ''

            results.append({
                'tarih': tarih.isoformat(),
                'gun': gun,
                'corba': corba,
                'ana_yemek': ana_yemek,
                'pilav_makarna': pilav,
                'tatli': tatli,
                'salata': salata,
                'ek_bilgi': ''
            })

    return results


# ═══ MAIN ═══
all_menus = []
print(f"Klasör: {FOLDER}\n")

for f in sorted(os.listdir(FOLDER)):
    if not f.endswith('.xlsx'):
        continue
    path = os.path.join(FOLDER, f)
    print(f"  📄 {f}")
    menus = parse_excel(path)
    print(f"     -> {len(menus)} gün menü")
    if menus:
        dates = [m['tarih'] for m in menus]
        print(f"     -> {min(dates)} ~ {max(dates)}")
    all_menus.extend(menus)

# Sırala ve tekrarları kaldır
all_menus.sort(key=lambda x: x['tarih'])
seen = set()
unique = []
for m in all_menus:
    if m['tarih'] not in seen:
        seen.add(m['tarih'])
        unique.append(m)

print(f"\n{'='*60}")
print(f"TOPLAM: {len(unique)} benzersiz gün")
if unique:
    print(f"Aralık: {unique[0]['tarih']} ~ {unique[-1]['tarih']}")

# Ay dağılımı
from collections import Counter
months = Counter(m['tarih'][:7] for m in unique)
print(f"\nAy dağılımı:")
for ay, n in sorted(months.items()):
    print(f"  {ay}: {n} gün")

# Nisan detay
print(f"\n📅 NİSAN 2026:")
for m in unique:
    if m['tarih'].startswith('2026-04'):
        print(f"  {m['tarih']} {m['gun']:10s} | {m['corba']:30s} | {m['ana_yemek']:45s} | {m['pilav_makarna']:25s} | T:{m['tatli']:20s} | S:{m['salata']}")

# Bugün
print(f"\n📅 BUGÜN (2026-04-10):")
bugun = [m for m in unique if m['tarih'] == '2026-04-10']
if bugun:
    m = bugun[0]
    print(f"  Çorba:     {m['corba']}")
    print(f"  Ana Yemek: {m['ana_yemek']}")
    print(f"  Pilav:     {m['pilav_makarna']}")
    print(f"  Tatlı:     {m['tatli']}")
    print(f"  Salata:    {m['salata']}")
else:
    print("  ❌ Bulunamadı!")

# CSV yaz
with open(OUTPUT, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['tarih', 'gun', 'corba', 'ana_yemek', 'pilav_makarna', 'tatli', 'salata', 'ek_bilgi'])
    writer.writeheader()
    writer.writerows(unique)

print(f"\n✅ CSV yazıldı: {OUTPUT} ({len(unique)} satır)")
